import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import time

# 1. ตั้งค่า Browser
options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.page_load_strategy = 'eager'
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

WAIT_TIMEOUT = int(os.getenv("EXPORT_WAIT_TIMEOUT", "60"))
PAGE_LOAD_TIMEOUT = int(os.getenv("EXPORT_PAGE_LOAD_TIMEOUT", "60"))
driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)

current_step = "เริ่มต้น"


def wait_for_any_condition(wait, conditions):
    last_error = None
    for condition in conditions:
        try:
            result = wait.until(condition)
            if result:
                return result
        except TimeoutException as exc:
            last_error = exc
    if last_error:
        raise last_error
    raise TimeoutException("ไม่พบเงื่อนไขที่คาดไว้: loginPage, app, หรือ table")


def normalize_name(value):
    text = str(value).strip()
    if not text:
        return ""

    # Keep only the left side for values like "aaa / bbb".
    if "/" in text:
        text = text.split("/", 1)[0].strip()

    tokens = " ".join(text.split()).split(" ")
    if len(tokens) >= 3 and tokens[0].lower() == tokens[-1].lower():
        tokens = tokens[:-1]

    return " ".join(tokens)


def format_excel_table(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    if ws.max_row < 2 or ws.max_column < 1:
        wb.save(file_path)
        return

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="ExportTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    wb.save(file_path)

try:
    current_step = "เปิดหน้าเว็บ"
    print("กำลังเปิดหน้าเว็บ...")
    url = "http://localhost:5173/app"
    driver.get(url)

    wait = WebDriverWait(driver, WAIT_TIMEOUT)

    # ถ้าโดน redirect ไปหน้า login ให้ล็อกอินก่อน
    if "loginPage" in driver.current_url:
        current_step = "รอหน้า login"
        email = os.getenv("NEON_EMAIL")
        password = os.getenv("NEON_PASSWORD")
        if not email or not password:
            raise RuntimeError(
                "ยังไม่ได้ล็อกอิน: Selenium ถูกพาไป /loginPage\n"
            )

        print("กำลังล็อกอิน...")
        current_step = "กรอกอีเมลและรหัสผ่าน"
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='email']"))).send_keys(email)
        driver.find_element(By.CSS_SELECTOR, "input[type='password']").send_keys(password)
        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

        # รอให้ระบบพากลับเข้า app หรืออย่างน้อยให้ login page หายไปก่อน
        current_step = "รอหลังล็อกอิน"
        wait_for_any_condition(
            wait,
            [
                EC.url_contains("/app"),
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "input[type='email']")),
            ],
        )

    # 2. รอจนตารางจาก Neon DB แสดงผล
    current_step = "รอหน้า app"
    print("กำลังรอให้ตารางโหลด...")
    wait.until(EC.url_contains("/app"))

    # บางครั้งข้อมูลเรนเดอร์ช้ากว่าการโหลดหน้า ให้ retry สั้น ๆ ก่อนยอมแพ้
    current_step = "รอตารางโหลด"
    table_loaded = False
    last_error = None
    for _ in range(3):
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr")))
            table_loaded = True
            break
        except TimeoutException as exc:
            last_error = exc
            time.sleep(2)

    if not table_loaded:
        raise TimeoutException(
            f"ติดตรง: {current_step}; รอโหลดตารางเกินเวลา {WAIT_TIMEOUT} วินาที; "
            f"current_url={driver.current_url}"
        ) from last_error

    # 3. อ่าน headers และ rows จาก DOM โดยตรง
    headers = [th.text.strip() for th in driver.find_elements(By.CSS_SELECTOR, "table thead th")]
    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    data = []
    for row in rows:
        cols = row.find_elements(By.CSS_SELECTOR, "td")
        values = [col.text.strip() for col in cols]
        if values:
            data.append(values)

    if not data:
        raise ValueError("ไม่พบข้อมูลในตาราง (table ว่าง)")

    # ป้องกันกรณีจำนวนคอลัมน์ไม่ตรง header
    max_cols = max(len(r) for r in data)
    if len(headers) < max_cols:
        headers.extend([f"col_{i}" for i in range(len(headers) + 1, max_cols + 1)])
    elif len(headers) > max_cols:
        headers = headers[:max_cols]

    normalized_data = [r + [""] * (max_cols - len(r)) for r in data]
    df = pd.DataFrame(normalized_data, columns=headers)

    # ลบคอลัมน์ที่ไม่จำเป็น
    removable_columns = ["Actions", "Column", ""]
    cols_to_drop = [c for c in removable_columns if c in df.columns]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)

    if 'Name' in df.columns:
        df['Name'] = df['Name'].map(normalize_name)

    output_file = "neon_data_export.xlsx"
    df.to_excel(output_file, index=False)
    format_excel_table(output_file)
    print(f"--- สำเร็จ! สร้างไฟล์ neon_data_export.xlsx จำนวน {len(df)} แถว ---")

except TimeoutException:
    print(f"เกิดข้อผิดพลาด: timeout ติดตรง {current_step}; current_url={driver.current_url}")

except Exception as e:
    print(f"เกิดข้อผิดพลาด: {e}")

finally:
    driver.quit()